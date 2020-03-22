###TEST CODE###

from flask import Flask,request,jsonify
import time
import bs4
import urllib.request
import key
import datetime
import schedule
import openpyxl as xl

import os
import sys


app = Flask(__name__)

#식당성정->날짜설정->아침점심저녁 설정->처음으로

Restaurant=["학생식당","푸름관","오름1동","오름3동","교직원 식당"]

ChoiceUrl=""
ChoiceDay=0
ChoiceRes=0

urlStudent="http://www.kumoh.ac.kr/ko/restaurant01.do"
urlProfess="http://www.kumoh.ac.kr/ko/restaurant02.do"
urlPorum="http://dorm.kumoh.ac.kr/dorm/restaurant_menu01.do"
urlorum1="http://dorm.kumoh.ac.kr/dorm/restaurant_menu02.do"
urlorum3="http://dorm.kumoh.ac.kr/dorm/restaurant_menu03.do"
urlBunsic="http://www.kumoh.ac.kr/ko/restaurant04.do"

urlGumidust="https://search.naver.com/search.naver?where=nexearch&sm=tab_etc&mra=blQ3&query=%EA%B2%BD%EB%B6%81%20%EB%AF%B8%EC%84%B8%EB%A8%BC%EC%A7%80"
urlGumiweather="http://www.kma.go.kr/wid/queryDFSRSS.jsp?zone=4719069000"
urlNaverGumiWeather = "https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&query=%EA%B5%AC%EB%AF%B8%EC%8B%9C+%EC%96%91%ED%8F%AC%EB%8F%99+%EB%82%A0%EC%94%A8&oquery=%EA%B5%AC%EB%AF%B8%EC%8B%9C+%EB%82%A0%EC%94%A8&tqi=UFk1%2BwprvxZssC9GFFdssssstU4-254477"

urlArr=[urlStudent,urlPorum,urlorum1,urlorum3,urlProfess,urlBunsic]
saveMenu = []  # 6개의 식당, 7개의 요일

'''
월요일~일요일 중식 : 0~6
월요일~일요일 석식 : 7~13

@@@ 예외적으로 오름 1동은 중식->조식 @@@
'''

jsonChoiceDay = {
    "version": "2.0",
    "template": {"outputs": [{"simpleText": {"text": "날짜를 선택해 주세요."}}],
                 "quickReplies": [{"label": "오늘", "action": "message", "messageText": "오늘"},
                                  {"label": "월", "action": "message", "messageText": "월"},
                                  {"label": "화", "action": "message", "messageText": "화"},
                                  {"label": "수", "action": "message", "messageText": "수"},
                                  {"label": "목", "action": "message", "messageText": "목"},
                                  {"label": "금", "action": "message", "messageText": "금"},
                                  {"label": "토", "action": "message", "messageText": "토"},
                                  {"label": "일", "action": "message", "messageText": "일"}
                                  ]
                 }
}

jsonChoiceRes = {
    "version": "2.0",
    "template": {"outputs": [{"simpleText": {"text": "식당을 선택해 주세요."}}],
                 "quickReplies": [{"label": "학생식당", "action": "message", "messageText": "학생식당"},
                                  {"label": "푸름관", "action": "message", "messageText": "푸름관"},
                                  {"label": "오름1동", "action": "message", "messageText": "오름1동"},
                                  {"label": "오름3동", "action": "message", "messageText": "오름3동"},
                                  {"label": "분식당", "action": "message", "messageText": "분식당"},
                                  {"label": "교직원", "action": "message", "messageText": "교직원"},
                                  ]
                 }
}

jsonChoiceAvailableTime = {
    "version": "2.0",
    "template": {"outputs": [{"simpleText": {"text": "식당을 선택해 주세요."}}],
                 "quickReplies": [{"label": "학생식당", "action": "message", "messageText": "학생식당 시간"},
                                  {"label": "기숙사", "action": "message", "messageText": "기숙사 시간"},
                                  {"label": "교직원", "action": "message", "messageText": "교직원 시간"},
                                  ]
                 }
}

StudentTime="조식시간 : 08:30 ~ 09:30\n중식시간 : 11:30 ~ 14:00\n석식시간 : 17:30 ~ 18:30\n토 : 10:00~14:00\n일,공휴일 : 휴무"
ProfessTime="중식시간 : 11:30 ~ 14:00\n석식시간 : 17:30 ~ 18:30"
DomitoryTime="학기중\n\n조식 시간\n- 평일 : 07:30 ~ 09:30\n- 주말 : 08:00 ~ 09:30\n중식 시간\n- 평일 : 11:30 ~ 13:30\n- 주말 : 12:00 ~ 13:30\n석식 시간\n- 평일 : 17:00 ~ 19:00\n- 주말 : 17:00 ~ 18:30\n\n방학중\n\n조식 시간- 08:00 ~ 09:30\n중식 시간- 12:00 ~ 13:30\n석식 시간- 17:00 ~ 18:30"



def returnMenu(url,num):  #식단 문자열을 반환하는 함수 (식당종류,날짜)

    global ChoiceRes

    html = bs4.BeautifulSoup(urllib.request.urlopen(url), "html.parser")
    menus=html.find("td")
    menu=str(menus.text)  #bs4 자료형을 String 형태로 변환, 식단의 존재 유무 판별

    if menu=="등록된 메뉴가 없습니다." : #식단이 없을경우(기숙사 식당 주로)
        return "등록된 메뉴가 없습니다. 😥"

    else:                              #식단이 있을경우
        menu = html.findAll("ul", {"class": "s-dot"})
        menuEnd = str(menu[num].text.rstrip("\n"))

        days=html.findAll("th",{"scope":{"col"}})
        day=str(days[num].text.lstrip())

        if menuEnd != "" :
            if ChoiceRes==2: #오름1동, 중식->조식
                menuEnd2 = str(menu[num + 7].text.rstrip("\n"))
                return "선택한 날짜 : "+day+"\n"+"아침메뉴\n\n"+menuEnd.lstrip()+"\n\n저녁메뉴\n\n"+menuEnd2.lstrip()

            elif ChoiceRes==5: #분식당, 1일 1메뉴
                return "선택한 날짜 : "+day+"\n"+menuEnd.lstrip()

            else:  #점심과 저녁
                menuEnd2 = str(menu[num + 7].text.rstrip("\n"))
                return "선택한 날짜 : "+day+"\n"+"점심메뉴\n\n"+menuEnd.lstrip()+"\n\n저녁메뉴\n\n"+menuEnd2.lstrip()

        else:
            return "등록된 메뉴가 없습니다. 😥"


def returnWeather(url):

    f = xl.Workbook()
    weatherxl = f.active

    html = bs4.BeautifulSoup(urllib.request.urlopen(url), "html.parser")
    weatherbox = html.find("div",{"class":"weather_area _mainArea"})

    today_weather = weatherbox.find("div",{"class":"info_data"})
    now_temp = today_weather.find("span",{"class":"todaytemp"})
    today_min_temp = today_weather.find("span",{"class":"min"})
    today_max_temp = today_weather.find("span",{"class":"max"})
    weatherxl.cell(1, 1, now_temp.text)
    weatherxl.cell(1, 2, today_min_temp.text)
    weatherxl.cell(1, 3, today_max_temp.text)

    #today_dust_box = weatherbox.find("dl",{"class":"indicator"})
    today_dusts = weatherbox.findAll("dd",{"class":"lv2"})
    today_parti_matter = weatherbox.find("dd",{"class":"lv1"})
    today_dust = today_dusts[0]
    today_ozon = today_dusts[1]
    weatherxl.cell(1, 4, today_parti_matter.text)
    weatherxl.cell(1, 5, today_dust.text)
    weatherxl.cell(1, 6, today_ozon.text)

    weather_predicts = weatherbox.findAll("li",{"class":{"date_info today"}})

    tom_weather = weather_predicts[1]
    tom_morning_rain = tom_weather.find("span",{"class":{"point_time morning"}})
    tom_afternoon_rain = tom_weather.find("span",{"class":{"point_time afternoon"}})
    tom_temp = tom_weather.find("dd")
    weatherxl.cell(2, 1, tom_morning_rain.text)
    weatherxl.cell(2, 2, tom_afternoon_rain.text)
    weatherxl.cell(2, 3, tom_temp.text)

    tom2_weather = weather_predicts[2]
    tom2_morning_rain = tom2_weather.find("span", {"class": {"point_time morning"}})
    tom2_afternoon_rain = tom2_weather.find("span", {"class": {"point_time afternoon"}})
    tom2_temp = tom2_weather.find("dd")
    weatherxl.cell(3, 1, tom2_morning_rain.text)
    weatherxl.cell(3, 2, tom2_afternoon_rain.text)
    weatherxl.cell(3, 3, tom2_temp.text)

    f.save('files/weather.xlsx')


returnWeather(urlNaverGumiWeather)


def returnWeatherjson(urlWeather,urlDust):

    f = xl.load_workbook('files/weather.xlsx', data_only=True)
    file = f['Sheet']

    temp = {
              "version": "2.0",
              "template": {
                "outputs": [
                  {
                    "carousel": {
                      "type": "basicCard",
                      "items": [
                        {
                          "title": "오늘 날씨",
                          "description":  f"현재 온도 : {file.cell(1,1).value}\n오늘 최저/최고 기온 : {file.cell(1,2).value}/{file.cell(1,3).value}\n"
                                          f"미세먼지 : {file.cell(1,4).value}\n초미세먼지 : {file.cell(1,5).value}\n오존 : {file.cell(1,6).value}"
                        },
                        {
                          "title": "내일 날씨",
                          "description": f"내일 최저/최고 기온 : {file.cell(2,3).value}\n내일 오전/오후 강수 확률 : {file.cell(2,1).value}/{file.cell(2,2).value}"
                        },
                        {
                          "title": "모레 날씨",
                          "description": f"모레 최저/최고 기온 : {file.cell(2,3).value}\n모레 오전/오후 강수 확률 : {file.cell(2,1).value}/{file.cell(2,2).value}"
                        }
                      ]
                    }
                  }
              ],
                  "quickReplies": [{"label": "처음으로", "action": "message", "messageText": "처음으로"}]
            }
    }

    return temp




ChoiceDay=0
ChoiceRes=4
emo="🌞⛅☔⚡⛄"

Restaurant=["학생식당","푸름관","오름1동","오름3동","교직원 식당","분식당"]


urlBustop = key.urlBustop
urlBusEnd = key.urlBusEnd

def secToMin(sec):
    return datetime.timedelta(seconds=sec)

def returnBus(url):

    html = bs4.BeautifulSoup(urllib.request.urlopen(url), "html.parser")
    body = html.find('body')
    print(type(body))

    totalCount = body.find('totalcount')
    totalCount = int(totalCount.text)

    arrtime = body.findAll('arrtime') #남은시간(초)
    routeid = body.findAll('routeid') #노선번호
    routeno = body.findAll('routeno') #버스번호

    for i in range(totalCount):
        print(str(routeno[i].text)+" "+str(routeid[i].text)+" "+str(secToMin(int(arrtime[i].text)))[2:])



def saveMenuArr():  #금오공대 전체 메뉴를 저장하기 위한 함수

    day = str(time.localtime().tm_mday)
    min = str(time.localtime().tm_min)
    sec = str(time.localtime().tm_sec)
    print(f"Save Start at {day} day, {min} min {sec} sec")

    f = xl.Workbook()
    menuxl = f.active

    global ChoiceRes
    global saveMenu
    a = -1
    ChoiceRes = 0

    for i in urlArr:   #식당 루프
        b = 0
        a += 1
        for j in range (7) :  #번호 루프
            menuxl.cell(a+1,b+1,returnMenu(i,j))  #해당하는 셀에 메뉴 정보를 저장
            b += 1
        ChoiceRes += 1

    f.save('files/data.xlsx')  #최종적으로 파일 저장

    min = str(time.localtime().tm_min)
    sec = str(time.localtime().tm_sec)
    print(f"Save Finish at {day} day {min} min {sec} sec")

#saveMenuArr()

#print(returnBus(urlBustop))
#print(returnDust(gumidust))
#print(returnBus(urlBustop))


