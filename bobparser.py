import time
import bs4
import urllib.request
import schedule
import openpyxl as xl


Restaurant=["학생식당","푸름관","오름1동","오름3동","교직원 식당"]

ChoiceUrl=""
ChoiceDay=0
ChoiceRes=0

urlStudent="http://www.kumoh.ac.kr/ko/restaurant01.do"
urlProfess="http://www.kumoh.ac.kr/ko/restaurant02.do"
urlPorum="http://dorm.kumoh.ac.kr/dorm/restaurant_menu01.do"
urlorum1="http://dorm.kumoh.ac.kr/dorm/restaurant_menu02.do"
urlorum3="http://dorm.kumoh.ac.kr/dorm/restaurant_menu03.do"
urlBunsic="http://www.kumoh.ac.kr/ko/restaurant04.do"

urlNaverGumiWeather = "https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&query=%EA%B5%AC%EB%AF%B8%EC%8B%9C+%EC%96%91%ED%8F%AC%EB%8F%99+%EB%82%A0%EC%94%A8&oquery=%EA%B5%AC%EB%AF%B8%EC%8B%9C+%EB%82%A0%EC%94%A8&tqi=UFk1%2BwprvxZssC9GFFdssssstU4-254477"

urlArr=[urlStudent,urlPorum,urlorum1,urlorum3,urlProfess,urlBunsic]


###변수 선언 완료
###함수 선언 시작


def returnMenu(url,num):  #식단 문자열을 반환하는 함수 (식당종류,날짜)

    global ChoiceRes

    html = bs4.BeautifulSoup(urllib.request.urlopen(url), "html.parser")
    menus=html.find("td")
    menu=str(menus.text)  #bs4 자료형을 String 형태로 변환, 식단의 존재 유무 판별

    if menu=="등록된 메뉴가 없습니다." : #식단이 없을경우(기숙사 식당 주로)
        return "등록된 메뉴가 없습니다. 😥"

    else:                              #식단이 있을경우
        menu = html.findAll("ul", {"class": "s-dot"})
        menuEnd = str(menu[num].text.rstrip("\n"))
        menuEnd = menuEnd.lstrip()

        days=html.findAll("th",{"scope":{"col"}})
        day=str(days[num].text.lstrip())

        if menuEnd != "" :
            if ChoiceRes==2:        #오름1동, 중식->조식
                menuEnd2 = str(menu[num + 7].text.rstrip("\n"))
                menuEnd2 = menuEnd2.lstrip()
                return f"선택한 날짜 : {day}\n아침메뉴\n\n{menuEnd}\n\n저녁메뉴\n\n{menuEnd2}"

            elif ChoiceRes==5:      #분식당, 1일 1메뉴
                return f"선택한 날짜 : {day}\n{menuEnd}"

            else:                    #점심과 저녁
                menuEnd2 = str(menu[num + 7].text.rstrip("\n"))
                menuEnd2 = menuEnd2.lstrip()
                return f"선택한 날짜 : {day}\n점심메뉴\n\n{menuEnd}\n\n저녁메뉴\n\n{menuEnd2}"

        else:
            return "등록된 메뉴가 없습니다. 😥"


def saveMenuArr():  #금오공대 전체 메뉴를 저장하기 위한 함수

    day = str(time.localtime().tm_mday)
    hour = str(time.localtime().tm_hour)
    min = str(time.localtime().tm_min)
    sec = str(time.localtime().tm_sec)
    print(f"Menu Save Start at {day} day, {hour}:{min}:{sec}")

    f = xl.Workbook()
    menuxl = f.active

    global ChoiceRes
    ChoiceRes = 0

    for i in urlArr:   #식당 루프
        b = 0
        for j in range (7) :  #번호 루프
            menuxl.cell(ChoiceRes+1,b+1,returnMenu(i,j))  #해당하는 셀에 메뉴 정보를 저장
            b += 1
        ChoiceRes += 10

    f.save('files/menu.xlsx')  #최종적으로 파일 저장

    hour = str(time.localtime().tm_hour)
    min = str(time.localtime().tm_min)
    sec = str(time.localtime().tm_sec)
    print(f"Menu Save Finish at {day} day,{hour}:{min}:{sec}")


def openMenu(a,b):  #해당 값의 셀 내용 반환하는 함수(x,y)

    f = xl.load_workbook('files/menu.xlsx',data_only=True)
    file = f['Sheet']

    return file.cell(a+1,b+1).value


def saveWeather(): #날씨 크롤링 후 엑셀에 저장하는 함수

    url = urlNaverGumiWeather

    day = str(time.localtime().tm_mday)
    hour = str(time.localtime().tm_hour)
    min = str(time.localtime().tm_min)
    sec = str(time.localtime().tm_sec)
    print(f"Weather Save Start at {day} day, {hour}:{min}:{sec}")

    f = xl.Workbook()
    weatherxl = f.active

    html = bs4.BeautifulSoup(urllib.request.urlopen(url), "html.parser")
    weatherbox = html.find("div",{"class":"weather_area _mainArea"})

    today_weather = weatherbox.find("div",{"class":"info_data"})
    now_temp = today_weather.find("span",{"class":"todaytemp"})
    today_min_temp = today_weather.find("span",{"class":"min"})
    today_max_temp = today_weather.find("span",{"class":"max"})
    weatherxl.cell(1, 1, now_temp.text)
    weatherxl.cell(1, 2, today_min_temp.text)
    weatherxl.cell(1, 3, today_max_temp.text)

    today_dust_box = weatherbox.find("dl",{"class":"indicator"})
    today_dusts = today_dust_box.findAll("dd")
    today_dust = today_dusts[0]
    today_parti_matter = today_dusts[1]
    today_ozon = today_dusts[2]
    weatherxl.cell(1, 4, today_dust.text)
    weatherxl.cell(1, 5, today_parti_matter.text)
    weatherxl.cell(1, 6, today_ozon.text)

    weather_predicts = weatherbox.findAll("li",{"class":{"date_info today"}})

    tom_weather = weather_predicts[1]
    tom_morning_rain = tom_weather.find("span",{"class":{"point_time morning"}})
    tom_morning_rain = tom_morning_rain.find("span", {"class": {"num"}})
    tom_afternoon_rain = tom_weather.find("span",{"class":{"point_time afternoon"}})
    tom_afternoon_rain = tom_afternoon_rain.find("span", {"class": {"num"}})
    tom_temp = tom_weather.find("dd")
    weatherxl.cell(2, 1, tom_morning_rain.text)
    weatherxl.cell(2, 2, tom_afternoon_rain.text)
    weatherxl.cell(2, 3, tom_temp.text)

    tom2_weather = weather_predicts[2]
    tom2_morning_rain = tom_weather.find("span",{"class":{"point_time morning"}})
    tom2_morning_rain = tom2_morning_rain.find("span", {"class": {"num"}})
    tom2_afternoon_rain = tom2_weather.find("span",{"class":{"point_time afternoon"}})
    tom2_afternoon_rain = tom2_afternoon_rain.find("span", {"class": {"num"}})
    tom2_temp = tom2_weather.find("dd")
    weatherxl.cell(3, 1, tom2_morning_rain.text)
    weatherxl.cell(3, 2, tom2_afternoon_rain.text)
    weatherxl.cell(3, 3, tom2_temp.text)

    f.save('files/weather.xlsx')

    hour = str(time.localtime().tm_hour)
    min = str(time.localtime().tm_min)
    sec = str(time.localtime().tm_sec)
    print(f"Weather Save Finish at {day} day,{hour}:{min}:{sec}")


def returnWeatherjson():  #날씨정보 엑셀에서 추출하는 함수

    f = xl.load_workbook('files/weather.xlsx', data_only=True)
    file = f['Sheet']

    temp = {
              "version": "2.0",
              "template": {
                "outputs": [
                  {
                    "carousel": {
                      "type": "basicCard",
                      "items": [
                        {
                          "title": "오늘 날씨",
                          "description":  f"현재 온도 : {file.cell(1,1).value}\n오늘 최저/최고 기온 : {file.cell(1,2).value}/{file.cell(1,3).value}\n"
                                          f"미세먼지 : {file.cell(1,4).value}\n초미세먼지 : {file.cell(1,5).value}\n오존 : {file.cell(1,6).value}"
                        },
                        {
                          "title": "내일 날씨",
                          "description": f"내일 최저/최고 기온 : {file.cell(2,3).value}\n내일 오전/오후 강수 확률 : {file.cell(2,1).value} % / {file.cell(2,2).value} %"
                        },
                        {
                          "title": "모레 날씨",
                          "description": f"모레 최저/최고 기온 : {file.cell(2,3).value}\n모레 오전/오후 강수 확률 : {file.cell(2,1).value} % / {file.cell(2,2).value} %"
                        }
                      ]
                    }
                  }
              ],
                  "quickReplies": [{"label": "처음으로", "action": "message", "messageText": "처음으로"}]
            }
    }

    return temp


saveMenuArr()  #프로그램 최초 실행 시 메뉴 리프레시(저장)
saveWeather()

schedule.every().monday.at("00:01").do(saveMenuArr)   #월요일 00:01 마다 크롤링,밑에도 알아서
schedule.every().monday.at("05:30").do(saveMenuArr)
schedule.every().wednesday.at("02:00").do(saveMenuArr)
schedule.every(15).minutes.do(saveWeather)

while True:
    schedule.run_pending()
    time.sleep(1)



