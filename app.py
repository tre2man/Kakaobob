from flask import Flask,request,jsonify
import time
import openpyxl as xl

import os
import sys

app = Flask(__name__)

ChoiceUrl=""
ChoiceWeek=0
ChoiceRes=0
Lastindex = 1

Restaurant=["학생식당","푸름관","오름1동","오름3동","교직원 식당","분식당"]
week=["월요일","화요일","수요일","목요일","금요일","토요일","일요일"]

StudentTime=str("조식시간 : 08:30 ~ 09:30\n중식시간 : 11:30 ~ 14:00\n석식시간 : 17:30 ~ 18:30\n토 : 10:00~14:00\n일,공휴일 : 휴무")

ProfessTime=str("중식시간 : 11:30 ~ 14:00\n석식시간 : 17:30 ~ 18:30")

DomitoryTime=str("학기중\n\n조식 시간\n- 평일 : 07:30 ~ 09:30\n- 주말 : 08:00 ~ 09:30\n중식 시간\n- 평일 : 11:30 ~ 13:30"
                 "\n- 주말 : 12:00 ~ 13:30\n석식 시간\n- 평일 : 17:00 ~ 19:00\n- 주말 : 17:00 ~ 18:30\n\n방학중\n\n"
                 "조식 시간- 08:00 ~ 09:30\n중식 시간- 12:00 ~ 13:30\n석식 시간- 17:00 ~ 18:30")

'''
월요일~일요일 중식 : 0~6
월요일~일요일 석식 : 7~13
@@@ 예외적으로 오름 1동은 중식->조식 @@@
'''

jsonMainmenu = {
    "version": "2.0",
    "template": {"outputs": [{"simpleText": {"text": "원하시는 기능을 선택해 주세요"}}],
                 "quickReplies": [{"label": "식단 정보", "action": "message", "messageText": "식단 정보"},
                                  {"label": "날씨 정보", "action": "message", "messageText": "날씨 정보"},
                                  {"label": "식당 이용 가능 시간", "action": "message", "messageText": "식당 이용 가능 시간"}
                                  ]
                 }
}

jsonChoiceRes = {
    "version": "2.0",
    "template": {"outputs": [{"simpleText": {"text": "🍽 식당을 선택해 주세요. 🍽"}}],
                 "quickReplies": [{"label": "학생식당", "action": "message", "messageText": "학생식당"},
                                  {"label": "푸름관", "action": "message", "messageText": "푸름관"},
                                  {"label": "오름1동", "action": "message", "messageText": "오름1동"},
                                  {"label": "오름3동", "action": "message", "messageText": "오름3동"},
                                  {"label": "교직원", "action": "message", "messageText": "교직원"},
                                  {"label": "분식당", "action": "message", "messageText": "분식당"}
                                  ]
                 }
}

jsonChoiceAvailableTime = {
    "version": "2.0",
    "template": {"outputs": [{"simpleText": {"text": "식당을 선택해 주세요."}}],
                 "quickReplies": [{"label": "학생식당", "action": "message", "messageText": "학생식당 시간"},
                                  {"label": "기숙사", "action": "message", "messageText": "기숙사 시간"},
                                  {"label": "교직원", "action": "message", "messageText": "교직원 시간"},
                                  ]
                 }
}


###변수 선언 완료
###함수 선언 시작


def returnMenujson(res,week):  #식당 메뉴를 json으로 리턴하는 함수

    f = xl.load_workbook('files/menu.xlsx', data_only=True)
    file = f['Sheet']

    temp = {
        "version": "2.0",
        "template": {"outputs": [{"simpleText": {"text": file.cell(res+1,week+1).value}}],
                     "quickReplies": [{"label": "처음으로", "action": "message", "messageText": "처음으로"},
                                      ]
                     }
        }

    return temp


def returnAvaliableTime(index):  #식당 이용 가능 시간을 json으로 리턴하는 함수

    temp = {
         "version": "2.0",
         "template": {
             "outputs": [{"simpleText": {"text": index}}],
             "quickReplies": [{"label": "처음으로", "action": "message", "messageText": "처음으로"},
                              ]
                     }
         }

    return temp


def returnjsonChoiceday():  #날짜 선택지를 json으로 리턴하는 함수

    temp = {
        "version": "2.0",
        "template": {"outputs": [{"simpleText": {
            "text": "📅 요일을 선택해 주세요. 📅\n\n오늘은 " + str(time.localtime().tm_year) + "년 " + str(
                time.localtime().tm_mon) + "월 " + str(time.localtime().tm_mday) + "일 " + week[
                        time.localtime().tm_wday] + " 입니다."}}],
                     "quickReplies": [{"label": "오늘", "action": "message", "messageText": "오늘"},
                                      {"label": "월요일", "action": "message", "messageText": "월요일"},
                                      {"label": "화요일", "action": "message", "messageText": "화요일"},
                                      {"label": "수요일", "action": "message", "messageText": "수요일"},
                                      {"label": "목요일", "action": "message", "messageText": "목요일"},
                                      {"label": "금요일", "action": "message", "messageText": "금요일"},
                                      {"label": "토요일", "action": "message", "messageText": "토요일"},
                                      {"label": "일요일", "action": "message", "messageText": "일요일"}
                                      ]
                     }
    }

    return temp


def returnWeatherjson():  #날씨 반환하는 함수

    f = xl.load_workbook('files/weather.xlsx', data_only=True)
    file = f['Sheet']

    temp = {
              "version": "2.0",
              "template": {
                "outputs": [
                  {
                    "carousel": {
                      "type": "basicCard",
                      "items": [
                        {
                          "title": "오늘 날씨",
                          "description":  f"현재 온도 : {file.cell(1,1).value}\n오늘 최저/최고 기온 : {file.cell(1,2).value}/{file.cell(1,3).value}\n"
                                          f"미세먼지 : {file.cell(1,4).value}\n초미세먼지 : {file.cell(1,5).value}\n오존 : {file.cell(1,6).value}"
                        },
                        {
                          "title": "내일 날씨",
                          "description": f"내일 최저/최고 기온 : {file.cell(2,3).value}\n내일 오전/오후 강수 확률 : {file.cell(2,1).value} % / {file.cell(2,2).value} %"
                        },
                        {
                          "title": "모레 날씨",
                          "description": f"모레 최저/최고 기온 : {file.cell(3,3).value}\n모레 오전/오후 강수 확률 : {file.cell(3,1).value} % / {file.cell(3,2).value} %"
                        }
                      ]
                    }
                  }
              ],
                  "quickReplies": [{"label": "처음으로", "action": "message", "messageText": "처음으로"}]
            }
    }

    return temp


def saveDBres(user,res):

    global Lastindex
    f = xl.load_workbook('files/user.xlsx', data_only=True)
    file = ['Sheet']

    for i in range(1, 501):
        if file.cell(i, 1).value == user :
            file.cell(i, 2, res)
            f.save('files/user.xlsx')
            print(f'Saved res in {i},{2}')
            return
        elif i == 500 :
            file.cell(Lastindex, 1, user)
            file.cell(Lastindex, 2, res)
            Lastindex += 1
            f.save('files/user.xlsx')
            print(f'Add user in {Lastindex},{1}')
            print(f'Saved res in {Lastindex},{2}')
            return


def findRes(user):

    f = xl.load_workbook('files/user.xlsx', data_only=True)
    file = f['Sheet']

    for i in range(1, 501):
        if file.cell(i, 1).value == user :
            return int(file.cell(i,2).value)


@app.route('/message', methods=['POST'])  #json으로 들어온 사용자 요청을 보고 판단
def bob():

    contents = request.get_json()    #사용자가 보낸 메세지 입력

    says = contents['userRequest']['utterance']         #사용자의 발화 추출
    user_key = contents['userRequest']['user']['id']    #사용자의 id 추출

    global ChoiceUrl,ChoiceRes,Choiceweek,jsonChoiceday,jsonChoiceRes

    if says==u"학생식당":
        response_data = returnjsonChoiceday()
        saveDBres(user_key, 0)

    elif says==u"푸름관":
        response_data = returnjsonChoiceday()
        saveDBres(user_key, 1)

    elif says==u"오름1동":
        response_data = returnjsonChoiceday()
        saveDBres(user_key, 2)

    elif says == u"오름3동":
        response_data = returnjsonChoiceday()
        saveDBres(user_key, 3)

    elif says==u"교직원":
        response_data = returnjsonChoiceday()
        saveDBres(user_key, 4)

    elif says==u"분식당":
        response_data = returnjsonChoiceday()
        saveDBres(user_key, 5)

    elif says==u"오늘":
        ChoiceWeek = time.localtime().tm_wday
        response_data = returnMenujson(findRes(user_key),ChoiceWeek)

    elif says==u"월요일":
        ChoiceWeek = 0
        response_data = returnMenujson(findRes(user_key),ChoiceWeek)

    elif says==u"화요일":
        ChoiceWeek = 1
        response_data = returnMenujson(findRes(user_key),ChoiceWeek)

    elif says==u"수요일":
        ChoiceWeek = 2
        response_data = returnMenujson(findRes(user_key),ChoiceWeek)

    elif says==u"목요일":
        ChoiceWeek = 3
        response_data = returnMenujson(findRes(user_key),ChoiceWeek)

    elif says==u"금요일":
        ChoiceWeek = 4
        response_data = returnMenujson(findRes(user_key),ChoiceWeek)

    elif says==u"토요일":
        ChoiceWeek = 5
        response_data = returnMenujson(findRes(user_key),ChoiceWeek)

    elif says==u"일요일":
        ChoiceWeek = 6
        response_data = returnMenujson(findRes(user_key),ChoiceWeek)

    elif says==u"처음으로" :
        response_data=jsonMainmenu

    elif says==u"식단 정보":
        response_data=jsonChoiceRes

    elif says==u"식당 이용 가능 시간":
        response_data=jsonChoiceAvailableTime

    elif says==u"학생식당 시간":
        response_data=returnAvaliableTime(StudentTime)

    elif says == u"기숙사 시간":
        response_data = returnAvaliableTime(DomitoryTime)

    elif says == u"교직원 시간":
        response_data = returnAvaliableTime(ProfessTime)

    elif says == u"날씨 정보":
        response_data = returnWeatherjson()

    else :
        response_data = jsonMainmenu

    return jsonify(response_data)

if __name__=="__main__":
     app.run(host="0.0.0.0", port=5000)