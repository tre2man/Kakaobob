import openpyxl as xl
import time

fopen = xl.load_workbook('files/user.xlsx', data_only=True)
f = fopen['Sheet']

#불러올 때 file.cell(2,3).value
def k():
    for i in range (1,500) :
        if f.cell(i,1).value == 'trees':
            print(i)
            f.cell(i,2,i)
            fopen.save('files/user.xlsx')
            return
        elif i == 500 :
            print(i)
            return

k()