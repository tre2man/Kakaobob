# 식단과 날씨 수정하는 부분

import time
import bs4
import urllib.request
import schedule
import openpyxl as xl

ChoiceRes = 0

urlStudent = "https://www.kumoh.ac.kr/ko/restaurant01.do"
urlProfess = "https://www.kumoh.ac.kr/ko/restaurant02.do"
urlPorum = "https://dorm.kumoh.ac.kr/dorm/restaurant_menu01.do"
urlorum1 = "https://dorm.kumoh.ac.kr/dorm/restaurant_menu02.do"
urlorum3 = "https://dorm.kumoh.ac.kr/dorm/restaurant_menu03.do"
urlBunsic = "https://www.kumoh.ac.kr/ko/restaurant04.do"

urlNaverGumiWeather = "https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&query=%" \
                      "EA%B5%AC%EB%AF%B8%EC%8B%9C+%EC%96%91%ED%8F%AC%EB%8F%99+%EB%82%A0%EC%94%A8&oquery" \
                      "=%EA%B5%AC%EB%AF%B8%EC%8B%9C+%EB%82%A0%EC%94%A8&tqi=UFk1%2BwprvxZssC9GFFdssssstU4-254477"

urlArr = [urlStudent, urlPorum, urlorum1, urlorum3, urlProfess, urlBunsic]


def returnMenu(url, num):  # 식단 문자열을 반환하는 함수 (식당종류,날짜)

    global ChoiceRes

    html = bs4.BeautifulSoup(urllib.request.urlopen(url), "html.parser")
    menus = html.find("td")
    menu = str(menus.text)  # bs4 자료형을 String 형태로 변환, 식단의 존재 유무 판별

    if menu == "등록된 메뉴가 없습니다.":  # 식단이 없을경우(기숙사 식당 주로)
        return "등록된 메뉴가 없습니다. 😥"

    else:  # 식단이 있을경우
        menu = html.findAll("ul", {"class": "s-dot"})
        menuEnd = str(menu[num].text.rstrip("\n"))
        menuEnd = menuEnd.lstrip()

        days = html.findAll("th", {"scope": {"col"}})
        day = str(days[num].text.lstrip())

        if menuEnd != "":

            # 각 식당마다 조식, 중식, 석식의 구성이 다르고 테이블이 나누어진 경우가 다르다.
            # 테이블이 나누어져 있으면 수식 실행, 나누어져 있지 않으면 실행하지 않는다.

            if ChoiceRes == 2:  # 오름1동 아침저녁 고정
                menuEnd2 = str(menu[num + 7].text.rstrip("\n"))
                menuEnd2 = menuEnd2.lstrip()
                return f"선택한 날짜 : {day}\n아침메뉴\n\n{menuEnd}\n\n저녁메뉴\n\n{menuEnd2}"

            else:
                try:
                    menuEnd2 = str(menu[num + 7].text.rstrip("\n"))
                    menuEnd2 = menuEnd2.lstrip()
                    return f"선택한 날짜 : {day}\n점심메뉴\n\n{menuEnd}\n\n저녁메뉴\n\n{menuEnd2}"

                except:
                    return f"선택한 날짜 : {day}\n오늘의식단\n\n{menuEnd}\n"

        else:
            return "등록된 메뉴가 없습니다. 😥"


def saveMenuArr():  # 금오공대 전체 메뉴를 엑셀에 저장하기 위한 함수

    day = str(time.localtime().tm_mday)
    hour = str(time.localtime().tm_hour)
    min = str(time.localtime().tm_min)
    sec = str(time.localtime().tm_sec)
    print(f"Menu Save Start at {day} day, {hour}:{min}:{sec}")

    try:
        f = xl.Workbook()
        menuxl = f.active

        global ChoiceRes
        ChoiceRes = 0

        for res in urlArr:  # 식당 루프
            col = 0
            for week in range(7):  # 번호 루프
                # 해당하는 셀에 메뉴 정보를 저장
                menuxl.cell(ChoiceRes + 1, col + 1, returnMenu(res, week))
                col += 1
            ChoiceRes += 1

        f.save('files/menu.xlsx')  # 최종적으로 파일 저장
        day = str(time.localtime().tm_mday)
        hour = str(time.localtime().tm_hour)
        min = str(time.localtime().tm_min)
        sec = str(time.localtime().tm_sec)
        print(f"Menu Save Finish at {day} day, {hour}:{min}:{sec}")

    except:
        print("Menu Save Error")


def saveWeather():  # 날씨 크롤링 후 엑셀에 저장하는 함수

    url = urlNaverGumiWeather
    url2 = urlTodayGumiWeather

    day = str(time.localtime().tm_mday)
    hour = str(time.localtime().tm_hour)
    min = str(time.localtime().tm_min)
    sec = str(time.localtime().tm_sec)
    print(f"Weather Save Start at {day} day, {hour}:{min}:{sec}")

    try:
        f = xl.Workbook()
        weatherxl = f.active

        html = bs4.BeautifulSoup(urllib.request.urlopen(url), "html.parser")

        weatherbox = html.find("div", {"class": "weather_area _mainArea"})

        today_weather = weatherbox.find("div", {"class": "info_data"})
        now_temp = today_weather.find("span", {"class": "todaytemp"})
        today_min_temp = today_weather.find("span", {"class": "min"})
        today_max_temp = today_weather.find("span", {"class": "max"})
        weatherxl.cell(1, 1, now_temp.text + "°")
        weatherxl.cell(1, 2, today_min_temp.text)
        weatherxl.cell(1, 3, today_max_temp.text)

        today_dust_box = weatherbox.find("dl", {"class": "indicator"})
        today_dusts = today_dust_box.findAll("dd")
        today_dust = today_dusts[0]
        today_parti_matter = today_dusts[1]
        today_ozon = today_dusts[2]
        weatherxl.cell(1, 4, today_dust.text)
        weatherxl.cell(1, 5, today_parti_matter.text)
        weatherxl.cell(1, 6, today_ozon.text)

        weather_predicts = weatherbox.findAll(
            "li", {"class": {"date_info today"}})

        tom_weather = weather_predicts[1]
        tom_morning_rain = tom_weather.find(
            "span", {"class": {"point_time morning"}})
        tom_morning_rain = tom_morning_rain.find("span", {"class": {"num"}})
        tom_afternoon_rain = tom_weather.find(
            "span", {"class": {"point_time afternoon"}})
        tom_afternoon_rain = tom_afternoon_rain.find(
            "span", {"class": {"num"}})
        tom_temp = tom_weather.find("dd")
        weatherxl.cell(2, 1, tom_morning_rain.text)
        weatherxl.cell(2, 2, tom_afternoon_rain.text)
        weatherxl.cell(2, 3, tom_temp.text)

        tom2_weather = weather_predicts[2]
        tom2_morning_rain = tom_weather.find(
            "span", {"class": {"point_time morning"}})
        tom2_morning_rain = tom2_morning_rain.find("span", {"class": {"num"}})
        tom2_afternoon_rain = tom2_weather.find(
            "span", {"class": {"point_time afternoon"}})
        tom2_afternoon_rain = tom2_afternoon_rain.find(
            "span", {"class": {"num"}})
        tom2_temp = tom2_weather.find("dd")
        weatherxl.cell(3, 1, tom2_morning_rain.text)
        weatherxl.cell(3, 2, tom2_afternoon_rain.text)
        weatherxl.cell(3, 3, tom2_temp.text)

        f.save('files/weather.xlsx')

        hour = str(time.localtime().tm_hour)
        min = str(time.localtime().tm_min)
        sec = str(time.localtime().tm_sec)
        print(f"Weather Save Finish at {day} day, {hour}:{min}:{sec}")
    except:
        print("Weather Save error")


saveMenuArr()  # 프로그램 최초 실행 시 메뉴 리프레시(저장)
saveWeather()

schedule.every().day.at("00:10").do(saveMenuArr)
schedule.every().day.at("00:20").do(saveMenuArr)
schedule.every().day.at("00:30").do(saveMenuArr)
schedule.every().day.at("06:00").do(saveMenuArr)
schedule.every().day.at("09:30").do(saveMenuArr)
schedule.every(15).minutes.do(saveWeather)

while True:
    schedule.run_pending()
    time.sleep(1)
