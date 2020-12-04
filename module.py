#길고 반복적인 변수 및 함수 부분

import bs4
import urllib.request
import time
import openpyxl as xl

Restaurant=["학생식당","푸름관","오름1동","오름3동","교직원 식당","분식당"]
week=["월요일","화요일","수요일","목요일","금요일","토요일","일요일"]

urlGumiBus = "http://211.236.110.97/GMBIS/m/page/srchBusArr.do?act=srchBusArr&stopId=132&stopKname=%EA%B8%88%EC%98%A4%EA%B3%B5%EB%8C%80%EC%A2%85%EC%A0%90&menuCode=1_03&stopServiceid=10132"

Lastindex = 1
user_max_number = 501    #저장 가능한 유저의 수

jsonMainmenu = {
    "version": "2.0",
    "template": {"outputs": [{"simpleText": {"text": "🔧 원하시는 기능을 선택해 주세요. 🔧"}}],
                 "quickReplies": [{"label": "식단 정보", "action": "message", "messageText": "식단 정보"},
                                  {"label": "버스 정보", "action": "message", "messageText": "버스 정보"},
                                  {"label": "날씨 정보", "action": "message", "messageText": "날씨 정보"},
                                  {"label": "식당 이용 가능 시간", "action": "message", "messageText": "식당 이용 가능 시간"},
                                  {"label": "정보", "action": "message", "messageText": "정보"}
                                  ]
                 }
}


jsonChoiceRes = {
    "version": "2.0",
    "template": {"outputs": [{"simpleText": {"text": "🍽 식당을 선택해 주세요. 🍽"}}],
                 "quickReplies": [{"label": "학생식당", "action": "message", "messageText": "학생식당"},
                                  {"label": "교직원", "action": "message", "messageText": "교직원"},
                                  {"label": "푸름관", "action": "message", "messageText": "푸름관"},
                                  {"label": "오름1동", "action": "message", "messageText": "오름1동"},
                                  {"label": "오름3동", "action": "message", "messageText": "오름3동"}
                                  ]
                 }
}


jsonChoiceAvailableTime = {
    "version": "2.0",
    "template": {"outputs": [{"simpleText": {"text": "식당을 선택해 주세요."}}],
                 "quickReplies": [{"label": "학생식당", "action": "message", "messageText": "학생식당 시간"},
                                  {"label": "기숙사", "action": "message", "messageText": "기숙사 시간"},
                                  {"label": "교직원", "action": "message", "messageText": "교직원 시간"},
                                  ]
                 }
}


def returnMenujson(res,week):  #식당 메뉴를 json으로 리턴하는 함수

    f = xl.load_workbook('files/menu.xlsx', data_only=True)
    file = f['Sheet']

    temp = {
        "version": "2.0",
        "template": {"outputs": [{"simpleText": {"text": file.cell(res+1,week+1).value}}],
                     "quickReplies": [{"label": "처음으로", "action": "message", "messageText": "처음으로"},
                                      ]
                     }
        }

    return temp


def returnAvaliableTime(index):  #식당 이용 가능 시간을 json으로 리턴하는 함수

    temp = {
         "version": "2.0",
         "template": {
             "outputs": [{"simpleText": {"text": index}}],
             "quickReplies": [{"label": "처음으로", "action": "message", "messageText": "처음으로"},
                              ]
                     }
         }

    return temp


def returnjsonChoiceday():  #날짜 선택지를 json으로 리턴하는 함수

    temp = {
        "version": "2.0",
        "template": {"outputs": [{"simpleText": {
            "text": "📅 요일을 선택해 주세요. 📅\n\n오늘은 " + str(time.localtime().tm_year) + "년 " + str(
                time.localtime().tm_mon) + "월 " + str(time.localtime().tm_mday) + "일 " + week[
                        time.localtime().tm_wday] + " 입니다."}}],
                     "quickReplies": [{"label": "오늘", "action": "message", "messageText": "오늘"},
                                      {"label": "월요일", "action": "message", "messageText": "월요일"},
                                      {"label": "화요일", "action": "message", "messageText": "화요일"},
                                      {"label": "수요일", "action": "message", "messageText": "수요일"},
                                      {"label": "목요일", "action": "message", "messageText": "목요일"},
                                      {"label": "금요일", "action": "message", "messageText": "금요일"},
                                      {"label": "토요일", "action": "message", "messageText": "토요일"},
                                      {"label": "일요일", "action": "message", "messageText": "일요일"}
                                      ]
                     }
    }

    return temp


def returnWeatherjson():  #날씨 데이터를 json으로 반환하는 함수

    f = xl.load_workbook('files/weather.xlsx', data_only=True)
    file = f['Sheet']

    temp = {
              "version": "2.0",
              "template": {
                "outputs": [
                  {
                    "carousel": {
                      "type": "basicCard",
                      "items": [
                        {
                          "title": "오늘 날씨",
                          "description":  f"현재 온도 : {file.cell(1,1).value}\n오늘 최저/최고 기온 : {file.cell(1,2).value}/{file.cell(1,3).value}\n"
                                          f"미세먼지 : {file.cell(1,4).value}\n초미세먼지 : {file.cell(1,5).value}\n오존 : {file.cell(1,6).value}"
                        },
                        {
                          "title": "내일 날씨",
                          "description": f"내일 최저/최고 기온 : {file.cell(2,3).value}\n내일 오전/오후 강수 확률 : {file.cell(2,1).value} % / {file.cell(2,2).value} %"
                        },
                        {
                          "title": "모레 날씨",
                          "description": f"모레 최저/최고 기온 : {file.cell(3,3).value}\n모레 오전/오후 강수 확률 : {file.cell(3,1).value} % / {file.cell(3,2).value} %"
                        }
                      ]
                    }
                  }
              ],
                  "quickReplies": [{"label": "처음으로", "action": "message", "messageText": "처음으로"}]
            }
    }

    return temp


def saveDBres(user,res):

    global Lastindex
    f = xl.load_workbook('files/user.xlsx', data_only=True)
    file = f['Sheet']

    for num in range(1, user_max_number):
        if file.cell(num, 1).value == user :
            file.cell(num, 2, res)
            f.save('files/user.xlsx')
            print(f'Saved res in {num}')

            return

        elif num == user_max_number-1 :
            file.cell(Lastindex, 1, user)
            file.cell(Lastindex, 2, res)
            f.save('files/user.xlsx')

            print(f'Add user in {Lastindex}')
            print(f'Saved res in {Lastindex}')

            Lastindex += 1

            return


def findRes(user):

    f = xl.load_workbook('files/user.xlsx', data_only=True)
    file = f['Sheet']

    for num in range(1, user_max_number):
        if file.cell(num, 1).value == user :
            return int(file.cell(num,2).value)


def returnBus():

    html = bs4.BeautifulSoup(urllib.request.urlopen(urlGumiBus), "html.parser")
    buses = html.findAll("ul",{"class":"arrive_desc"})
    value = ""
    if buses == "":
        return "버스가 없습니다."
    else:
        for bus in buses:
            bus_no = bus.find("li",{"class":"bus_no"}).text
            bus_state = bus.find("li",{"class":"bus_state"}).text
            bus_now = bus.findAll("li")
            value += f"\n{bus_no} {bus_state} {bus_now[3].text}"

        return value.lstrip("\n")


def returnBusTime():

    temp = {
        "version": "2.0",
        "template": {
            "outputs": [
                {
                    "carousel": {
                        "type": "basicCard",
                        "items": [
                            {
                                "title": "금오공대 종점 정류장\n버스 번호 / 남은 시간 / 현재 위치",
                                "description": returnBus()
                            }
                        ]
                    }
                }
            ],
            "quickReplies": [{"label": "처음으로", "action": "message", "messageText": "처음으로"}]
        }
    }

    return temp
